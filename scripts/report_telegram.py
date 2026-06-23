#!/usr/bin/env python3
"""Build an XLSX sanity report from Allure results and deliver it to Telegram.

Reads ``allure-results/*-result.json`` (the human-readable case titles live
there, e.g. ``"7. Интеграция с 1С — данные импортируются"``), produces an Excel
report mirroring BusinessHub_Sanity_Report3.xlsx (№ | Кейс | Статус), and posts
a short summary + the file to a Telegram chat.

Usage:
    python scripts/report_telegram.py [--results DIR] [--out FILE] [--no-send]
    python scripts/report_telegram.py --chat-id     # resolve TG_CHAT_ID helper

Config comes from .env via config.settings (TG_BOT_TOKEN, TG_CHAT_ID,
TG_REPORT_ENV). If the token/chat are unset, the report file is still written
but the send step is skipped — so wiring this into CI is safe before creds exist.
"""

from __future__ import annotations

import argparse
import html
import json
import re
import sys
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Allow ``python scripts/report_telegram.py`` from the repo root.
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from config.settings import settings  # noqa: E402

TZ = ZoneInfo("Asia/Tashkent")

# Allure status → (russian label, fill colour, counts-as-failure).
STATUS_META = {
    "passed": ("passed", "C6EFCE", False),
    "failed": ("failed", "FFC7CE", True),
    "broken": ("broken", "FFC7CE", True),
    "skipped": ("skipped", "FFEB9C", False),
    "unknown": ("unknown", "D9D9D9", False),
}

_NUM_RE = re.compile(r"^\s*(\d+)")


def _case_num(name: str) -> int:
    """Leading case number from an Allure title; 999 if absent (sorts last)."""
    m = _NUM_RE.match(name or "")
    return int(m.group(1)) if m else 999


def load_results(results_dir: Path) -> list[dict]:
    """Collapse Allure result JSONs into one row per case.

    With ``--reruns`` a flaky test yields several result files for the same
    title; keep the final attempt (latest ``stop`` timestamp) so the report
    reflects the outcome that decided the run.
    """
    latest: dict[str, dict] = {}
    for f in sorted(results_dir.glob("*-result.json")):
        try:
            data = json.loads(f.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, OSError):
            continue
        name = data.get("name")
        if not name:
            continue
        stop = data.get("stop", 0)
        prev = latest.get(name)
        if prev is None or stop >= prev.get("stop", 0):
            latest[name] = data

    rows = []
    for data in latest.values():
        name = data["name"]
        status = data.get("status", "unknown")
        detail = (data.get("statusDetails") or {}).get("message", "") or ""
        rows.append(
            {
                "num": _case_num(name),
                "name": name,
                "status": status if status in STATUS_META else "unknown",
                "detail": detail.strip().splitlines()[0] if detail.strip() else "",
            }
        )
    rows.sort(key=lambda r: r["num"])
    return rows


def summarize(rows: list[dict]) -> dict[str, int]:
    counts = {"passed": 0, "failed": 0, "broken": 0, "skipped": 0, "unknown": 0}
    for r in rows:
        counts[r["status"]] += 1
    counts["total"] = len(rows)
    return counts


def build_xlsx(rows: list[dict], counts: dict, out_path: Path, when: datetime) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sanity"

    title_font = Font(bold=True, size=13)
    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="305496")
    center = Alignment(horizontal="center", vertical="center")
    wrap = Alignment(vertical="center", wrap_text=True)

    # Title + summary banner.
    ws["A1"] = f"MyXodim Sanity — {settings.tg_report_env} — {when:%Y-%m-%d %H:%M %Z}"
    ws["A1"].font = title_font
    ws.merge_cells("A1:D1")
    ws["A2"] = (
        f"✅ {counts['passed']} passed   "
        f"❌ {counts['failed'] + counts['broken']} failed   "
        f"⏭ {counts['skipped']} skipped   "
        f"(всего {counts['total']})"
    )
    ws.merge_cells("A2:D2")

    # Header row.
    headers = ["№", "Кейс", "Статус", "Примечание"]
    hrow = 4
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=hrow, column=c, value=h)
        cell.font = head_font
        cell.fill = head_fill
        cell.alignment = center

    # Data rows.
    for i, r in enumerate(rows, start=hrow + 1):
        label, colour, _ = STATUS_META[r["status"]]
        # Strip the leading "N. " from the title — number lives in its own column.
        case_text = re.sub(r"^\s*\d+\.\s*", "", r["name"])
        ws.cell(row=i, column=1, value=r["num"]).alignment = center
        ws.cell(row=i, column=2, value=case_text).alignment = wrap
        st = ws.cell(row=i, column=3, value=label)
        st.alignment = center
        st.fill = PatternFill("solid", fgColor=colour)
        ws.cell(row=i, column=4, value=r["detail"]).alignment = wrap

    widths = {1: 6, 2: 60, 3: 12, 4: 50}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A5"

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def build_caption(counts: dict, rows: list[dict], when: datetime) -> str:
    failed = counts["failed"] + counts["broken"]
    head = "✅ ВСЕ ЗЕЛЁНЫЕ" if failed == 0 else f"❌ ЕСТЬ ПАДЕНИЯ ({failed})"
    lines = [
        f"<b>MyXodim Sanity</b> — {settings.tg_report_env}",
        f"{when:%Y-%m-%d %H:%M %Z}",
        "",
        head,
        f"✅ {counts['passed']}   ❌ {failed}   ⏭ {counts['skipped']}   "
        f"(всего {counts['total']})",
    ]
    bad = [r for r in rows if STATUS_META[r["status"]][2]]
    if bad:
        lines.append("")
        lines.append("<b>Упавшие кейсы:</b>")
        for r in bad:
            note = f" — {html.escape(r['detail'])}" if r["detail"] else ""
            lines.append(f"❌ {html.escape(r['name'])}{note}")
    return "\n".join(lines)


def tg_api(method: str) -> str:
    return f"https://api.telegram.org/bot{settings.tg_bot_token}/{method}"


def resolve_chat_id() -> int:
    """Print recent chats the bot can see (run after messaging the bot/group)."""
    if not settings.tg_bot_token:
        print("TG_BOT_TOKEN is empty — set it in .env first.", file=sys.stderr)
        return 2
    r = requests.get(tg_api("getUpdates"), timeout=30)
    r.raise_for_status()
    seen = {}
    for upd in r.json().get("result", []):
        msg = upd.get("message") or upd.get("channel_post") or {}
        chat = msg.get("chat")
        if chat:
            seen[chat["id"]] = chat.get("title") or chat.get("username") or chat.get("first_name", "")
    if not seen:
        print("No chats found. Send a message to the bot (or add it to the "
              "group and post there), then re-run.", file=sys.stderr)
        return 1
    print("Chats visible to the bot — copy the id into TG_CHAT_ID:")
    for cid, label in seen.items():
        print(f"  {cid}\t{label}")
    return 0


def send(xlsx_path: Path, caption: str) -> bool:
    if not (settings.tg_bot_token and settings.tg_chat_id):
        print("Telegram send skipped (TG_BOT_TOKEN/TG_CHAT_ID unset).")
        return False
    with xlsx_path.open("rb") as fh:
        resp = requests.post(
            tg_api("sendDocument"),
            data={
                "chat_id": settings.tg_chat_id,
                "caption": caption,
                "parse_mode": "HTML",
            },
            files={"document": (xlsx_path.name, fh,
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            timeout=60,
        )
    if not resp.ok:
        print(f"Telegram send failed: {resp.status_code} {resp.text}", file=sys.stderr)
        return False
    print(f"Telegram report sent to chat {settings.tg_chat_id}.")
    return True


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--results", default="allure-results", help="Allure results dir")
    ap.add_argument("--out", default=None, help="XLSX output path")
    ap.add_argument("--no-send", action="store_true", help="build file, don't post to TG")
    ap.add_argument("--chat-id", action="store_true",
                    help="list chats the bot can see (helper to find TG_CHAT_ID)")
    args = ap.parse_args()

    if args.chat_id:
        return resolve_chat_id()

    results_dir = Path(args.results)
    if not results_dir.is_dir():
        print(f"Results dir not found: {results_dir}", file=sys.stderr)
        return 2
    rows = load_results(results_dir)
    if not rows:
        print(f"No *-result.json found in {results_dir}", file=sys.stderr)
        return 2

    when = datetime.now(TZ)
    counts = summarize(rows)
    out_path = Path(args.out) if args.out else Path(
        "test-results") / f"sanity-report-{when:%Y%m%d-%H%M%S}.xlsx"
    build_xlsx(rows, counts, out_path, when)
    print(f"Report written: {out_path}  "
          f"(✅{counts['passed']} ❌{counts['failed'] + counts['broken']} "
          f"⏭{counts['skipped']})")

    if not args.no_send:
        send(out_path, build_caption(counts, rows, when))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
